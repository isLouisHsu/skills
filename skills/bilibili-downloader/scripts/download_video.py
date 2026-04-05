#!/usr/bin/env python3
"""
B站视频下载脚本
支持：单个视频下载、专辑批量下载

用法：
    python download_video.py single <bvid> [output_dir]
    python download_video.py album <mid> <season_id> [output_dir]
"""

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Dict, List, Optional
from urllib.parse import urlparse, parse_qs

import av
import requests
from tenacity import retry, stop_after_attempt, wait_exponential

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class VideoInfo:
    """视频信息数据类"""
    bvid: str = ''
    avid: int = 0
    title: str = ''
    desc: str = ''
    duration: str = ''
    duration_sec: int = 0
    pubdate: int = 0
    pub_date_str: str = ''
    play_count: int = 0
    danmaku_count: int = 0
    like_count: int = 0
    coin_count: int = 0
    share_count: int = 0
    favorite_count: int = 0
    reply_count: int = 0
    cover_url: str = ''
    video_link: str = ''
    cover_path: str = ''
    video_path: str = ''


class BilibiliCrawler:
    """B站爬虫类"""

    def __init__(self, delay: float = 1.5, max_retries: int = 3):
        """
        初始化爬虫
        :param delay: 请求间隔秒数（反爬）
        :param max_retries: 网络异常时的最大重试次数
        """
        self.delay = delay
        self.max_retries = max_retries
        self.session = requests.Session()

        # 设置请求头（模拟浏览器）
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://www.bilibili.com/',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }

    @retry(stop=stop_after_attempt(4), wait=wait_exponential(multiplier=1, min=1, max=8))
    def _fetch(self, url: str, params: Optional[Dict] = None, headers: Optional[Dict] = None) -> requests.Response:
        """带重试的请求方法"""
        merged_headers = {**self.headers, **(headers or {})}
        response = self.session.get(url, params=params, headers=merged_headers, timeout=30)
        response.raise_for_status()
        return response

    def _format_duration(self, seconds: int) -> str:
        """将秒数格式化为 mm:ss 或 hh:mm:ss"""
        if not seconds:
            return "00:00"
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{secs:02d}"
        return f"{minutes:02d}:{secs:02d}"

    def _format_timestamp(self, timestamp: int) -> str:
        """将时间戳格式化为可读字符串"""
        if not timestamp:
            return ''
        return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(timestamp))

    def _download_cover(self, cover_url: str, save_path: str) -> Optional[str]:
        """
        下载封面图片
        :param cover_url: 封面URL
        :param save_path: 保存路径
        :return: 本地路径或None
        """
        if not cover_url:
            return None

        if os.path.exists(save_path):
            return save_path

        time.sleep(self.delay)

        for attempt in range(self.max_retries + 1):
            try:
                resp = self._fetch(cover_url)
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                with open(save_path, 'wb') as f:
                    f.write(resp.content)
                return save_path
            except Exception as e:
                if attempt == self.max_retries:
                    print(f"  封面下载失败: {e}")
                    return None
                wait_time = 2 ** attempt
                print(f"  封面下载失败，{wait_time}秒后重试...")
                time.sleep(wait_time)

    def _stream_download(self, url: str, save_path: str, label: str, referer: str) -> None:
        """流式下载单个URL到文件，带进度输出"""
        download_headers = {**self.headers, 'Referer': referer}
        with self.session.get(url, headers=download_headers, stream=True, timeout=60) as resp:
            resp.raise_for_status()
            total = int(resp.headers.get('content-length', 0))
            downloaded = 0
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            with open(save_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total:
                            pct = downloaded / total * 100
                            print(f"\r  {label}: {pct:.1f}% ({downloaded // 1024 // 1024}MB / {total // 1024 // 1024}MB)", end='', flush=True)
        print()

    def _remux(self, video_path: str, audio_path: str, output_path: str) -> None:
        """用 PyAV 将视频流和音频流 remux 为 MP4，不重新编码"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with av.open(video_path) as v_in, av.open(audio_path) as a_in, \
                av.open(output_path, 'w', format='mp4') as out:
            v_in_stream = v_in.streams.video[0]
            a_in_stream = a_in.streams.audio[0]
            v_out = out.add_stream_from_template(v_in_stream)
            a_out = out.add_stream_from_template(a_in_stream)
            for packet in v_in.demux(v_in_stream):
                if packet.dts is None or packet.size == 0:
                    continue
                packet.stream = v_out
                out.mux(packet)
            for packet in a_in.demux(a_in_stream):
                if packet.dts is None or packet.size == 0:
                    continue
                packet.stream = a_out
                out.mux(packet)

    def download_video(self, bvid: str, cid: int, save_path: str) -> Optional[str]:
        """
        下载视频MP4文件（DASH格式，视频+音频分别下载后合并）
        :param bvid: 视频BV号
        :param cid: 视频cid
        :param save_path: 保存路径
        :return: 本地路径或None
        """
        if os.path.exists(save_path):
            print(f"  视频已存在，跳过: {save_path}")
            return save_path

        url = 'https://api.bilibili.com/x/player/playurl'
        params = {
            'bvid': bvid,
            'cid': cid,
            'qn': 80,
            'fnval': 16,
            'fnver': 0,
            'fourk': 0,
        }
        referer = f'https://www.bilibili.com/video/{bvid}'

        tmp_video = tmp_audio = None
        try:
            response = self._fetch(url, params=params, headers={'Referer': referer})
            data = response.json()

            if data.get('code') != 0:
                print(f"  获取视频流失败: {data.get('message', '未知错误')}")
                return None

            dash = data.get('data', {}).get('dash')
            if not dash:
                print("  未获取到 DASH 流信息")
                return None

            video_url = dash['video'][0]['baseUrl']
            audio_url = dash['audio'][0]['baseUrl']

            save_dir = os.path.dirname(save_path) or '.'
            tmp_video = os.path.join(save_dir, f"_{bvid}_video.m4s")
            tmp_audio = os.path.join(save_dir, f"_{bvid}_audio.m4s")

            print(f"  下载视频流: {bvid}...")
            self._stream_download(video_url, tmp_video, '视频流', referer)
            print(f"  下载音频流: {bvid}...")
            self._stream_download(audio_url, tmp_audio, '音频流', referer)

            print(f"  合并音视频...")
            self._remux(tmp_video, tmp_audio, save_path)

            print(f"  视频已保存: {save_path}")
            return save_path

        except Exception as e:
            print(f"  视频下载失败: {e}")
            if os.path.exists(save_path):
                os.remove(save_path)
            return None
        finally:
            for tmp in (tmp_video, tmp_audio):
                if tmp and os.path.exists(tmp):
                    os.remove(tmp)

    def _parse_video_data(self, video: Dict) -> VideoInfo:
        """解析视频原始数据为VideoInfo对象"""
        info = VideoInfo()
        info.bvid = video.get('bvid', '')
        info.avid = video.get('avid', 0)
        info.title = video.get('title', '')
        info.desc = video.get('desc', '')
        info.duration_sec = video.get('duration', 0)
        info.duration = self._format_duration(info.duration_sec)
        info.pubdate = video.get('pubdate', 0)
        info.pub_date_str = self._format_timestamp(info.pubdate)

        stat = video.get('stat', {})
        info.play_count = stat.get('view', 0)
        info.danmaku_count = stat.get('danmaku', 0)
        info.like_count = stat.get('like', 0)
        info.coin_count = stat.get('coin', 0)
        info.share_count = stat.get('share', 0)
        info.favorite_count = stat.get('favorite', 0)
        info.reply_count = stat.get('reply', 0)

        info.cover_url = video.get('pic', '')
        info.video_link = f"https://www.bilibili.com/video/{info.bvid}"

        return info

    def get_cid(self, bvid: str) -> int:
        """通过 BV 号获取视频 cid"""
        url = 'https://api.bilibili.com/x/web-interface/view'
        resp = self._fetch(
            url, params={'bvid': bvid},
            headers={'Referer': f'https://www.bilibili.com/video/{bvid}'}
        )
        return resp.json().get('data', {}).get('cid', 0)

    def get_video_info(self, bvid: str) -> Optional[VideoInfo]:
        """获取单个视频的详细信息"""
        url = 'https://api.bilibili.com/x/web-interface/view'
        params = {'bvid': bvid}

        try:
            response = self._fetch(url, params=params, headers={'Referer': f'https://www.bilibili.com/video/{bvid}'})
            data = response.json()

            if data.get('code') != 0:
                print(f"API错误: {data.get('message', '未知错误')}")
                return None

            video_data = data.get('data', {})
            return self._parse_video_data(video_data)
        except Exception as e:
            print(f"获取视频详情失败: {e}")
            return None

    def fetch_album_videos(self, mid: int, season_id: int) -> tuple:
        """
        拉取专辑所有视频列表，返回 (season_title, videos)
        :param mid: 用户ID
        :param season_id: 专辑ID
        :return: (season_title, videos)
        """
        videos = []
        page_num = 1
        season_title = f"season_{season_id}"

        while True:
            print(f"获取第 {page_num} 页视频列表...")
            url = 'https://api.bilibili.com/x/polymer/web-space/seasons_archives_list'
            params = {
                'mid': mid,
                'season_id': season_id,
                'sort_reverse': 'false',
                'page_num': page_num,
                'page_size': 30,
            }
            resp = self._fetch(
                url, params=params,
                headers={'Referer': f'https://space.bilibili.com/{mid}/'}
            )
            data = resp.json()

            if data.get('code') != 0:
                print(f"API错误: {data.get('message', '未知错误')}")
                break

            d = data['data']
            meta = d.get('meta', {})
            season_title = meta.get('name', season_title)
            archives = d.get('archives', [])

            for v in archives:
                videos.append(self._parse_video_data(v))

            page_info = d.get('page', {})
            total = page_info.get('total', 0)
            page_size = page_info.get('page_size', 30)
            total_pages = (total + page_size - 1) // page_size

            print(f"  本页 {len(archives)} 个，累计 {len(videos)} / {total} 个")

            if page_num >= total_pages or not archives:
                break

            page_num += 1
            time.sleep(self.delay)

        return season_title, videos

    def export_to_excel(self, videos: List[VideoInfo], output_path: str, sheet_name: str = "视频列表") -> Optional[str]:
        """导出视频列表到Excel"""
        if not HAS_OPENPYXL:
            print("警告: 未安装 openpyxl，跳过Excel导出")
            return None

        if not videos:
            print("没有视频数据可导出")
            return None

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = [
            '序号', 'BV号', '标题', '描述', '时长', '发布时间',
            '播放量', '弹幕数', '点赞数', '投币数', '收藏数', '分享数', '评论数',
            '视频链接', '封面本地路径', '视频本地路径', '封面URL'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row, video in enumerate(videos, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=video.bvid)
            ws.cell(row=row, column=3, value=video.title)
            ws.cell(row=row, column=4, value=video.desc)
            ws.cell(row=row, column=5, value=video.duration)
            ws.cell(row=row, column=6, value=video.pub_date_str)
            ws.cell(row=row, column=7, value=video.play_count)
            ws.cell(row=row, column=8, value=video.danmaku_count)
            ws.cell(row=row, column=9, value=video.like_count)
            ws.cell(row=row, column=10, value=video.coin_count)
            ws.cell(row=row, column=11, value=video.favorite_count)
            ws.cell(row=row, column=12, value=video.share_count)
            ws.cell(row=row, column=13, value=video.reply_count)
            ws.cell(row=row, column=14, value=video.video_link)
            ws.cell(row=row, column=15, value=video.cover_path)
            ws.cell(row=row, column=16, value=video.video_path)
            ws.cell(row=row, column=17, value=video.cover_url)

        column_widths = {1: 6, 2: 12, 3: 50, 4: 40, 5: 10, 6: 20,
                         7: 12, 8: 10, 9: 10, 10: 10, 11: 10, 12: 10, 13: 10,
                         14: 35, 15: 40, 16: 40, 17: 50}
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        for row in range(2, len(videos) + 2):
            for col in [3, 4, 14, 15, 16, 17]:
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        ws.freeze_panes = 'A2'
        wb.save(output_path)
        return output_path

    def download_single(self, bvid: str, output_dir: str) -> Optional[VideoInfo]:
        """下载单个视频"""
        print(f"开始获取视频详情: {bvid}")

        video = self.get_video_info(bvid)

        if not video:
            print(f"获取视频信息失败: {bvid}")
            return None

        # 使用 {标题}-{bvid} 格式创建目录
        safe_title = re.sub(r'[\\/:*?"<>|]', '_', video.title)[:50]
        video_dir = os.path.join(output_dir, f"{safe_title}-{bvid}")
        os.makedirs(video_dir, exist_ok=True)

        # 下载封面
        if video.cover_url:
            ext = video.cover_url.split('.')[-1].split('@')[0]
            ext = ext if ext in ['jpg', 'jpeg', 'png', 'webp'] else 'jpg'
            cover_path = os.path.join(video_dir, f"cover.{ext}")
            print(f"  下载封面: {video.title[:30]}...")
            video.cover_path = self._download_cover(video.cover_url, cover_path) or ''

        # 获取cid并下载视频
        try:
            cid = self.get_cid(bvid)
            if cid:
                time.sleep(self.delay)
                video_path = os.path.join(video_dir, "video.mp4")
                video.video_path = self.download_video(bvid, cid, video_path) or ''
        except Exception as e:
            print(f"  下载视频失败: {e}")

        print(f"\n视频信息:")
        print(f"  标题: {video.title}")
        print(f"  时长: {video.duration}")
        print(f"  播放量: {video.play_count}")
        print(f"  发布时间: {video.pub_date_str}")

        # 导出Excel
        if HAS_OPENPYXL:
            excel_path = os.path.join(video_dir, "info.xlsx")
            self.export_to_excel([video], excel_path, sheet_name="视频详情")
            print(f"\nExcel已保存: {excel_path}")

        return video

    def download_album(self, mid: int, season_id: int, output_dir: str) -> List[VideoInfo]:
        """下载专辑视频列表（按视频分文件夹存储封面和视频）"""
        print(f"开始爬取专辑: mid={mid}, season_id={season_id}")
        print("-" * 50)

        season_title, videos = self.fetch_album_videos(mid, season_id)

        # 创建专辑文件夹：{专辑名称}-{season_id}
        safe_season_title = re.sub(r'[\\/:*?"<>|]', '_', season_title)
        album_dir = os.path.join(output_dir, f"album/{safe_season_title}-{season_id}")
        os.makedirs(album_dir, exist_ok=True)
        print(f"输出目录: {album_dir}")

        print("\n" + "=" * 50)
        print(f"爬取完成！专辑: {season_title}")
        print(f"共获取 {len(videos)} 个视频")
        print("=" * 50)

        # 逐个处理视频：复用 download_single 方法
        for idx, video in enumerate(videos, 1):
            print(f"\n[{idx}/{len(videos)}] 处理视频: {video.title}")

            # 复用 download_single 方法，它会自动创建 {标题}-{bvid} 子目录
            self.download_single(video.bvid, album_dir)

        # 导出Excel到专辑根目录
        if videos and HAS_OPENPYXL:
            safe_title = re.sub(r'[\\/:*?"<>|]', '_', season_title)
            excel_path = os.path.join(album_dir, f"{safe_title}.xlsx")
            self.export_to_excel(videos, excel_path)
            print(f"\nExcel已保存: {excel_path}")

        print("\n" + "=" * 50)
        print(f"专辑下载完成: {album_dir}")
        print("目录结构:")
        album_folder_name = os.path.basename(album_dir)
        print(f"  {album_folder_name}/")
        print(f"    ├── {safe_season_title}.xlsx")
        for v in videos:
            safe_v_title = re.sub(r'[\\/:*?"<>|]', '_', v.title)[:50]
            print(f"    ├── {safe_v_title}-{v.bvid}/")
            print(f"    │   ├── cover.*")
            print(f"    │   └── video.mp4")
        print("=" * 50)

        return videos


def extract_ids_from_url(url: str) -> tuple:
    """从B站URL中提取MID和SEASON_ID"""
    parsed = urlparse(url)
    params = parse_qs(parsed.query)

    mid = None
    season_id = None

    # 尝试从URL参数中提取
    if 'mid' in params:
        mid = int(params['mid'][0])
    if 'season_id' in params:
        season_id = int(params['season_id'][0])

    return mid, season_id


def main():
    parser = argparse.ArgumentParser(description='B站视频下载器')
    subparsers = parser.add_subparsers(dest='command', help='子命令')

    # single 命令
    single_parser = subparsers.add_parser('single', help='下载单个视频')
    single_parser.add_argument('bvid', help='视频BV号')
    single_parser.add_argument('output_dir', nargs='?', default='./downloads', help='输出目录')

    # album 命令
    album_parser = subparsers.add_parser('album', help='下载专辑视频列表')
    album_parser.add_argument('mid', help='UP主MID（或完整URL）')
    album_parser.add_argument('season_id', nargs='?', help='专辑SEASON_ID（如果从URL解析可省略）')
    album_parser.add_argument('output_dir', nargs='?', default='./downloads', help='输出目录')

    args = parser.parse_args()

    # 创建爬虫实例
    crawler = BilibiliCrawler(delay=1.5)

    if args.command == 'single':
        crawler.download_single(args.bvid, args.output_dir)
    elif args.command == 'album':
        # 如果第一个参数是URL，尝试解析
        mid = args.mid
        if mid.startswith('http'):
            parsed_mid, parsed_season_id = extract_ids_from_url(mid)
            if parsed_mid and parsed_season_id:
                mid = parsed_mid
                season_id = parsed_season_id
            else:
                print("无法从URL解析MID和SEASON_ID，请手动提供")
                sys.exit(1)
        else:
            mid = int(mid)
            season_id = int(args.season_id) if args.season_id else None

        if not season_id:
            print("需要提供SEASON_ID")
            sys.exit(1)

        crawler.download_album(mid, season_id, args.output_dir)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
